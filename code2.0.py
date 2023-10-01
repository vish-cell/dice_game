import cv2 
from datetime import date
import face_recognition as fr
from tkinter import Tk
import os
import numpy as np
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook 
import openpyxl as op
import pandas as pd
import datetime

def encode_all_faces(folder):    
    list_people_encoding = []
    
    for filename in os.listdir(folder):
        single_person_file=[]  
        for image in os.listdir("{}\{}".format(folder,filename)):
            full_path = os.path.join(filename,image)
            known_image = fr.load_image_file("{}\{}".format(folder,full_path))
            known_encoding = fr.face_encodings(known_image)
            known_face=np.array(known_encoding)
            single_person_file.append((known_face))
        list_people_encoding.append((single_person_file, filename))
    return list_people_encoding

def find_target_face(target_image):
    people=[]
    face_locations = fr.face_locations(target_image)
    target_encodings = fr.face_encodings(target_image)
    print("targets length.....",len(face_locations))

    list1=encode_all_faces("load_image2.0")
    for arrays,file in encode_faces:
        for array in arrays:
            for target_encode in target_encodings:
                unknown_face=np.array(target_encode)
                is_target_face = fr.compare_faces([array][0], unknown_face, tolerance=0.50)
                if is_target_face==[]:
                    continue
                else:
                    if is_target_face[0]:
                        if file not in people:
                            people.append(file)
    print("recognized faces=",len(people))
    print("unrecognized face=",len(face_locations)-len(people))     
    return people
            





def main():
    while True:
        ret, frame = camera.read()  # Reading camera input

        if not ret:
            print("Failed to grab frame")
            break
        cv2.imshow("Camera", frame)

        k = cv2.waitKey(1)
        if k % 256 == 32:
            image_name = "image_{}.png".format(date.today())
            cv2.imwrite(image_name, frame)
            print(image_name + " has been taken")
            return image_name




camera = cv2.VideoCapture(0)  #webcamera opens
load_image = main()
camera.release()
cv2.destroyAllWindows()

Tk().withdraw()
target_image =  fr.load_image_file(askopenfilename())  #fr.load_image_file(load_image) askopenfilename()
encode_faces=encode_all_faces("load_image2.0")
present=find_target_face(target_image)
print(present)



# uploading data into excel sheets
wb = op.load_workbook('excel sheets/sheet_1.xlsx')                   #give the full path of the file here
sheet = wb.active
rows=[]
columns=1
date1=datetime.date.strftime(date.today(), "%y-%m-%d")


for i in range(2,27):
    cellref=sheet.cell(row=1, column=i)
    if cellref.value !=None:
        col_date=datetime.date.strftime(cellref.value, "%y-%m-%d")
        if col_date==date1:
            columns=i
            break
        

for i in range(3,15):
    cellref=sheet.cell(row=i, column=1)
    for j in present:
        if cellref.value== j:
            rows.append(i)
print(rows)
print(columns)

for i in rows:
    cellref=sheet.cell(row=i,column=columns)
    cellref.value='p'

for x in range(3,15):
    if x not in rows:
        cellref=sheet.cell(row=x,column=columns)
        cellref.value='a'


wb.save('excel sheets/sheet_1.xlsx')
wb.close()